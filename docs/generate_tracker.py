"""
Generate Excel task tracker for Dashboard / BE / Mobile
Features: Product, Diskon, Promosi
"""
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY

OUTPUT = "Task Tracker - Phase 2.xlsx"

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY       = "1F497D"
BLUE_MED   = "2E74B5"
BLUE_LIGHT = "BDD7EE"
BLUE_PALE  = "DEEAF1"

PLAT_COLOR = {
    "BE":        ("E2EFDA", "375623"),   # green bg / dark green text
    "Dashboard": ("FFF2CC", "7F6000"),   # yellow bg / dark yellow text
    "Mobile":    ("FCE4D6", "833C00"),   # orange bg / dark orange text
}

STATUS_COLOR = {
    "TODO":        "F2F2F2",
    "IN PROGRESS": "FFF2CC",
    "DONE":        "E2EFDA",
    "BLOCKED":     "FCE4D6",
    "REVIEW":      "EAD1DC",
}

PRIO_COLOR = {
    "High":   "FF0000",
    "Medium": "ED7D31",
    "Low":    "70AD47",
}

thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_BORDER = Border(
    left=Side(border_style="medium", color=NAVY),
    right=Side(border_style="medium", color=NAVY),
    top=Side(border_style="medium", color=NAVY),
    bottom=Side(border_style="medium", color=NAVY),
)


def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def hfont(hex_color, bold=False, size=10):
    return Font(color=hex_color, bold=bold, size=size, name="Calibri")


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Task data ─────────────────────────────────────────────────────────────────
# (Feature, Platform, Category, Task, Priority)
TASKS = [
    # ── PRODUCT ──────────────────────────────────────────────────────────────
    # BE
    ("Product", "BE", "CRUD",     "CRUD Produk SIMPLE (nama, harga, SKU, kategori, gambar)", "High"),
    ("Product", "BE", "CRUD",     "CRUD Produk VARIANT — ProductVariantGroup + ProductVariant", "High"),
    ("Product", "BE", "CRUD",     "CRUD Produk MODIFIER — ProductModifierGroup + ProductModifier", "High"),
    ("Product", "BE", "CRUD",     "API upload & simpan gambar produk (Images)", "Medium"),
    ("Product", "BE", "Business", "Validasi perubahan productType (blokir revert jika ada transaksi)", "High"),
    ("Product", "BE", "Business", "Migrasi tipe: hapus variant/modifier saat ganti ke SIMPLE (safe delete)", "Medium"),
    ("Product", "BE", "Stock",    "Tambah kolom variantId (nullable) ke tabel stock & StockMovement", "High"),
    ("Product", "BE", "Stock",    "Validasi stock per variant saat transaksi (VARIANT product)", "High"),
    ("Product", "BE", "Stock",    "Validasi stock di level produk saat transaksi (MODIFIER product)", "Medium"),
    ("Product", "BE", "API",      "GET /products — list dengan filter kategori, tipe, isActive, search", "High"),
    ("Product", "BE", "API",      "GET /products/{id} — detail + variant/modifier groups", "High"),
    # Dashboard
    ("Product", "Dashboard", "UI", "Halaman list produk — filter, search, tipe badge", "High"),
    ("Product", "Dashboard", "UI", "Form tambah/edit produk SIMPLE", "High"),
    ("Product", "Dashboard", "UI", "Form tambah/edit produk VARIANT — manajemen group & item variant", "High"),
    ("Product", "Dashboard", "UI", "Form tambah/edit produk MODIFIER — manajemen group & item modifier", "High"),
    ("Product", "Dashboard", "UI", "Upload foto produk dari form", "Medium"),
    ("Product", "Dashboard", "UI", "Konfirmasi & blokir perubahan tipe produk jika ada transaksi", "Medium"),
    ("Product", "Dashboard", "UI", "Manajemen stok per variant (input qty per variant item)", "High"),
    # Mobile
    ("Product", "Mobile", "Kasir", "Tampil katalog produk di layar kasir (grid/list)", "High"),
    ("Product", "Mobile", "Kasir", "Search produk by nama / SKU", "High"),
    ("Product", "Mobile", "Kasir", "Pilih variant — modal pilihan variant group (wajib jika isRequired)", "High"),
    ("Product", "Mobile", "Kasir", "Pilih modifier — modal pilihan modifier group (min/max select)", "High"),
    ("Product", "Mobile", "Kasir", "Tampil harga modifier sebagai addon di cart item", "Medium"),
    ("Product", "Mobile", "Kasir", "Blokir tambah item jika stok = 0", "High"),

    # ── DISKON ───────────────────────────────────────────────────────────────
    # BE
    ("Diskon", "BE", "CRUD",     "CRUD Diskon (nama, tipe, nilai, scope, kode, tanggal berlaku)", "High"),
    ("Diskon", "BE", "Business", "Validasi kode diskon — aktif, belum kadaluarsa, limit pemakaian", "High"),
    ("Diskon", "BE", "Business", "Hitung diskon scope ALL (dari total subtotal)", "High"),
    ("Diskon", "BE", "Business", "Hitung diskon scope PRODUCT (per item yang cocok productId)", "High"),
    ("Diskon", "BE", "Business", "Hitung diskon scope CATEGORY (per item yang masuk kategori)", "High"),
    ("Diskon", "BE", "Business", "Validasi maksimum 1 diskon per transaksi", "High"),
    ("Diskon", "BE", "Business", "Endpoint GET /discounts/active — list diskon codeless (tanpa kode)", "Medium"),
    ("Diskon", "BE", "Business", "Increment usageCount setelah transaksi PAID", "Medium"),
    # Dashboard
    ("Diskon", "Dashboard", "UI", "Halaman list diskon — status badge, tanggal berlaku", "High"),
    ("Diskon", "Dashboard", "UI", "Form tambah/edit diskon — tipe PERCENTAGE / NOMINAL", "High"),
    ("Diskon", "Dashboard", "UI", "Konfigurasi scope: ALL / PRODUCT (pilih produk) / CATEGORY (pilih kategori)", "High"),
    ("Diskon", "Dashboard", "UI", "Toggle aktif/nonaktif diskon", "Medium"),
    ("Diskon", "Dashboard", "UI", "Tampilkan usage count & limit di list", "Low"),
    # Mobile
    ("Diskon", "Mobile", "Kasir", "Input kode diskon di halaman cart/payment", "High"),
    ("Diskon", "Mobile", "Kasir", "Pilih diskon dari list (codeless) — sheet bottom dengan list aktif", "High"),
    ("Diskon", "Mobile", "Kasir", "Tampil potongan diskon per item dan total di cart", "High"),
    ("Diskon", "Mobile", "Kasir", "Hapus / ganti diskon yang sudah dipilih", "Medium"),
    ("Diskon", "Mobile", "Kasir", "Pesan error jika kode tidak valid / expired / limit habis", "Medium"),

    # ── PROMOSI ───────────────────────────────────────────────────────────────
    # BE
    ("Promosi", "BE", "CRUD",     "CRUD Promosi (nama, tipe, kondisi, tanggal, priority, canCombine)", "High"),
    ("Promosi", "BE", "Business", "Auto-apply tipe DISCOUNT_BY_ORDER — diskon dari subtotal order", "High"),
    ("Promosi", "BE", "Business", "Auto-apply tipe BUY_X_GET_Y — deteksi item qualifying & gratis", "High"),
    ("Promosi", "BE", "Business", "BUY_X_GET_Y: isMultiplied — kalikan reward jika qty > threshold", "Medium"),
    ("Promosi", "BE", "Business", "Auto-apply tipe DISCOUNT_BY_ITEM_SUBTOTAL — diskon per item jika subtotal item >= N", "High"),
    ("Promosi", "BE", "Business", "Logika priority — urutkan promosi, terapkan tertinggi dulu", "High"),
    ("Promosi", "BE", "Business", "Logika canCombine — terapkan semua jika true, berhenti di canCombine=false tertinggi", "High"),
    ("Promosi", "BE", "Business", "Filter promosi per outlet (outletId) saat kalkulasi", "Medium"),
    ("Promosi", "BE", "API",      "GET /promotions/applicable — list promosi yang akan diterapkan untuk cart", "Medium"),
    # Dashboard
    ("Promosi", "Dashboard", "UI", "Halaman list promosi — tipe badge, status aktif, priority", "High"),
    ("Promosi", "Dashboard", "UI", "Form tambah/edit promosi — tipe DISCOUNT_BY_ORDER", "High"),
    ("Promosi", "Dashboard", "UI", "Form tambah/edit promosi — tipe BUY_X_GET_Y (pilih item X, item Y gratis)", "High"),
    ("Promosi", "Dashboard", "UI", "Form tambah/edit promosi — tipe DISCOUNT_BY_ITEM_SUBTOTAL", "High"),
    ("Promosi", "Dashboard", "UI", "Konfigurasi priority & canCombine", "Medium"),
    ("Promosi", "Dashboard", "UI", "Assign promosi ke outlet tertentu", "Medium"),
    # Mobile
    ("Promosi", "Mobile", "Kasir", "Tampil badge/label promosi yang aktif diterapkan di cart", "High"),
    ("Promosi", "Mobile", "Kasir", "Tampil item gratis hasil BUY_X_GET_Y di cart (harga Rp 0)", "High"),
    ("Promosi", "Mobile", "Kasir", "Rincian potongan promosi di halaman payment summary", "High"),
    ("Promosi", "Mobile", "Kasir", "Informasi promo aktif hari ini di halaman kasir (banner / chip)", "Low"),
]

COLUMNS = [
    ("No",         5),
    ("Feature",    12),
    ("Platform",   13),
    ("Category",   13),
    ("Task",       52),
    ("Priority",   11),
    ("Status",     14),
    ("Assignee",   16),
    ("Est. Days",  11),
    ("Start Date", 13),
    ("End Date",   13),
    ("Notes",      30),
]


def build_header(ws):
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = hfill(NAVY)
        cell.font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
        cell.alignment = center()
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def build_tracker_sheet(ws):
    ws.title = "Tracker"
    build_header(ws)

    # Data validations
    # We'll set sqref after writing all rows; collect cells first
    status_cells = []
    prio_cells = []

    row = 2
    prev_feature = None
    counter = 1

    for (feature, platform, category, task, priority) in TASKS:
        # Feature group separator row
        if feature != prev_feature:
            if prev_feature is not None:
                row += 1  # blank spacer
            # Group header
            ws.row_dimensions[row].height = 18
            gc = ws.cell(row=row, column=1, value=feature.upper())
            gc.fill = hfill(BLUE_MED)
            gc.font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
            gc.alignment = left()
            gc.border = BORDER
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
            row += 1
            prev_feature = feature
            counter = 1

        bg, fg = PLAT_COLOR[platform]

        values = [
            counter,
            feature,
            platform,
            category,
            task,
            priority,
            "TODO",  # default status
            "",      # assignee
            "",      # est days
            "",      # start
            "",      # end
            "",      # notes
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = left() if col_idx == 5 else center()
            cell.font = Font(name="Calibri", size=10)

            # Platform color on Platform column
            if col_idx == 3:
                cell.fill = hfill(bg)
                cell.font = Font(name="Calibri", size=10, bold=True, color=fg)
            # Priority color
            elif col_idx == 6:
                prio_hex = PRIO_COLOR.get(priority, "000000")
                cell.font = Font(name="Calibri", size=10, bold=True, color=prio_hex)
            # Status default style
            elif col_idx == 7:
                cell.fill = hfill(STATUS_COLOR["TODO"])
            else:
                cell.fill = hfill("FFFFFF")

            # Collect for validation
            if col_idx == 7:
                status_cells.append(cell.coordinate)
            if col_idx == 6:
                prio_cells.append(cell.coordinate)

        ws.row_dimensions[row].height = 30
        row += 1
        counter += 1

    # Register data validations
    if status_cells:
        status_dv = DataValidation(
            type="list",
            formula1='"TODO,IN PROGRESS,DONE,BLOCKED,REVIEW"',
            allow_blank=True,
            showDropDown=False,
            sqref=" ".join(status_cells),
        )
        ws.add_data_validation(status_dv)
    if prio_cells:
        prio_dv = DataValidation(
            type="list",
            formula1='"High,Medium,Low"',
            allow_blank=True,
            showDropDown=False,
            sqref=" ".join(prio_cells),
        )
        ws.add_data_validation(prio_dv)


def build_summary_sheet(ws, wb):
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "Task Tracker — CASHPOS Phase 2"
    title.fill = hfill(NAVY)
    title.font = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
    title.alignment = center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = "Feature: Product  |  Diskon  |  Promosi       Platform: Dashboard  |  BE  |  Mobile"
    sub.fill = hfill(BLUE_LIGHT)
    sub.font = Font(color=NAVY, bold=False, size=10, name="Calibri")
    sub.alignment = center()
    ws.row_dimensions[2].height = 18

    # Count tasks per feature & platform from TASKS list
    from collections import Counter
    counts = Counter((f, p) for f, p, *_ in TASKS)
    features = ["Product", "Diskon", "Promosi"]
    platforms = ["BE", "Dashboard", "Mobile"]

    # Header
    r = 4
    headers = ["Feature", "BE", "Dashboard", "Mobile", "Total", ""]
    col_widths = [14, 12, 14, 12, 10, 1]
    for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = hfill(BLUE_MED)
        cell.font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
        cell.alignment = center()
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[r].height = 20
    r += 1

    totals = {p: 0 for p in platforms}
    grand_total = 0
    for feat in features:
        row_vals = [feat]
        row_total = 0
        for plat in platforms:
            n = counts.get((feat, plat), 0)
            row_vals.append(n)
            row_total += n
            totals[plat] += n
        row_vals.append(row_total)
        grand_total += row_total
        for c, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = hfill(BLUE_PALE) if c == 1 else hfill("FFFFFF")
            cell.font = Font(name="Calibri", size=10, bold=(c == 1))
            cell.alignment = center()
            cell.border = BORDER
        ws.row_dimensions[r].height = 18
        r += 1

    # Totals row
    total_row = ["TOTAL"] + [totals[p] for p in platforms] + [grand_total]
    for c, v in enumerate(total_row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = hfill(NAVY)
        cell.font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
        cell.alignment = center()
        cell.border = HEADER_BORDER
    ws.row_dimensions[r].height = 20
    r += 2

    # Legend — Platform
    ws.merge_cells(f"A{r}:B{r}")
    ws.cell(row=r, column=1, value="Platform Legend").fill = hfill(BLUE_MED)
    ws.cell(row=r, column=1).font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    ws.cell(row=r, column=1).alignment = center()
    r += 1
    for plat, (bg, fg) in PLAT_COLOR.items():
        ws.cell(row=r, column=1, value=plat).fill = hfill(bg)
        ws.cell(row=r, column=1).font = Font(color=fg, bold=True, size=10, name="Calibri")
        ws.cell(row=r, column=1).alignment = center()
        ws.cell(row=r, column=1).border = BORDER
        r += 1
    r += 1

    # Legend — Status
    ws.merge_cells(f"A{r}:B{r}")
    ws.cell(row=r, column=1, value="Status Legend").fill = hfill(BLUE_MED)
    ws.cell(row=r, column=1).font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    ws.cell(row=r, column=1).alignment = center()
    r += 1
    for status, bg in STATUS_COLOR.items():
        ws.cell(row=r, column=1, value=status).fill = hfill(bg)
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).alignment = center()
        ws.cell(row=r, column=1).border = BORDER
        r += 1
    r += 1

    # Legend — Priority
    ws.merge_cells(f"A{r}:B{r}")
    ws.cell(row=r, column=1, value="Priority Legend").fill = hfill(BLUE_MED)
    ws.cell(row=r, column=1).font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    ws.cell(row=r, column=1).alignment = center()
    r += 1
    for prio, color in PRIO_COLOR.items():
        ws.cell(row=r, column=1, value=prio).fill = hfill("FFFFFF")
        ws.cell(row=r, column=1).font = Font(color=color, bold=True, size=10, name="Calibri")
        ws.cell(row=r, column=1).alignment = center()
        ws.cell(row=r, column=1).border = BORDER
        r += 1


def main():
    wb = Workbook()
    ws_tracker = wb.active
    ws_summary = wb.create_sheet()

    build_tracker_sheet(ws_tracker)
    build_summary_sheet(ws_summary, wb)

    # Move summary to first position
    wb.move_sheet("Summary", offset=-wb.index(wb["Summary"]))

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")
    print(f"Total tasks: {len(TASKS)}")
    from collections import Counter
    by_feature = Counter(f for f, *_ in TASKS)
    by_platform = Counter(p for _, p, *_ in TASKS)
    for k, v in sorted(by_feature.items()):
        print(f"  {k}: {v} tasks")
    for k, v in sorted(by_platform.items()):
        print(f"  {k}: {v} tasks")


if __name__ == "__main__":
    main()
